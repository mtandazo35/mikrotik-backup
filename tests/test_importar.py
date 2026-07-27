"""Prueba del alta masiva: lo que sale de Excel tiene que entrar sin pelearse.

No toca la red ni el disco del inventario: todo son bytes en memoria, igual que
lo que llega por una subida HTTP desde el panel.

Lo que se comprueba es que el archivo se DECODIFICA bien (separador, BOM,
codificacion, orden y sinonimos de la cabecera). La validacion de cada equipo
es cosa de inventory.validar_equipo y se prueba aparte; aqui solo se verifica
que las filas llegan enteras hasta ella.

Tambien se comprueban las columnas opcionales 'usuario' y 'clave': vacias es lo
normal (se heredan las credenciales generales) y la plantilla tiene que avisar
de que llenarlas convierte el archivo en una lista de contrasenas en claro.

Ejecutar:  python -m tests.test_importar
"""

import io

from mkbackup.importar import (
    COLUMNAS,
    ErrorImportacion,
    hay_soporte_xlsx,
    leer,
    plantilla_csv,
    plantilla_xlsx,
)
from mkbackup.inventory import validar_equipo

FALLOS = []


def comprobar(descripcion: str, condicion: bool) -> None:
    print(f"[{'OK  ' if condicion else 'FALLA'}] {descripcion}")
    if not condicion:
        FALLOS.append(descripcion)


def lanza(descripcion: str, funcion) -> None:
    """Comprueba que algo levanta ErrorImportacion y no otra cosa."""
    try:
        funcion()
    except ErrorImportacion:
        comprobar(descripcion, True)
    except Exception as exc:  # noqa: BLE001 - queremos verlo si pasa
        comprobar(f"{descripcion} (salio {type(exc).__name__}: {exc})", False)
    else:
        comprobar(f"{descripcion} (no lanzo nada)", False)


def main() -> None:
    # 1. CSV con coma, el formato canonico.
    csv_coma = (
        "nombre,empresa,ip,puerto,grupo\r\n"
        "BTS-Cuenca-01,Andinanet,10.20.1.1,22,bts\r\n"
        "BTS-Cuenca-02,Andinanet,10.20.1.2,,bts\r\n"
    ).encode("utf-8")
    filas, avisos = leer(csv_coma, "equipos.csv")
    comprobar("CSV con coma: lee las dos filas", len(filas) == 2)
    comprobar("CSV con coma: no genera avisos", avisos == [])
    comprobar("CSV con coma: la primera fila de datos es la 2",
              filas[0].numero == 2)
    comprobar("CSV con coma: todas las columnas presentes",
              set(filas[0].datos) == set(COLUMNAS))
    comprobar("CSV con coma: los valores llegan enteros",
              filas[0].datos == {"nombre": "BTS-Cuenca-01", "empresa": "Andinanet",
                                 "ip": "10.20.1.1", "puerto": "22", "grupo": "bts",
                                 "intervalo_minutos": "", "usuario": "",
                                 "clave": ""})
    comprobar("CSV con coma: el puerto vacio se queda vacio, no se inventa",
              filas[1].datos["puerto"] == "")
    comprobar("CSV sin columna intervalo: la clave existe y viene vacia",
              filas[0].datos["intervalo_minutos"] == "")
    comprobar("CSV sin columnas usuario/clave: existen y vienen vacias "
              "(se heredan las credenciales generales)",
              filas[0].datos["usuario"] == "" and filas[0].datos["clave"] == "")

    # 2. CSV con punto y coma: lo que produce Excel en configuracion regional
    # espanola, que es la de las maquinas donde se llena la plantilla.
    csv_puntoycoma = (
        "nombre;empresa;ip;puerto;grupo\r\n"
        "NODO-Loja;Fibra Austral;190.15.72.10;2222;core\r\n"
    ).encode("utf-8")
    filas, _ = leer(csv_puntoycoma, "equipos.csv")
    comprobar("CSV con punto y coma: detecta el separador", len(filas) == 1)
    comprobar("CSV con punto y coma: no parte mal los campos",
              filas[0].datos["ip"] == "190.15.72.10"
              and filas[0].datos["puerto"] == "2222")

    # 3. BOM y CRLF: exactamente lo que escribe "Guardar como CSV UTF-8".
    con_bom = b"\xef\xbb\xbf" + (
        "nombre,empresa,ip,puerto,grupo\r\n"
        "AP-Machala,Andinanet,10.30.0.5,22,aps\r\n"
    ).encode("utf-8")
    filas, avisos = leer(con_bom, "equipos.csv")
    comprobar("CSV con BOM: el BOM no se pega a la primera cabecera",
              len(filas) == 1 and filas[0].datos["nombre"] == "AP-Machala")
    comprobar("CSV con CRLF: no queda retorno de carro en la ultima columna",
              filas[0].datos["grupo"] == "aps")
    comprobar("CSV con BOM: no avisa de codificacion", avisos == [])

    # 4. cp1252 con enye: Excel en Windows guarda asi si no eliges UTF-8.
    csv_cp1252 = (
        "nombre,empresa,ip,puerto,grupo\r\n"
        "RB-Ninez,Compania Andina de Nuevas Redes,10.40.0.1,22,core\r\n"
    ).replace("Ninez", "Niñez").replace("Compania", "Compañía")
    filas, avisos = leer(csv_cp1252.encode("cp1252"), "equipos.csv")
    comprobar("CSV cp1252: se lee igualmente", len(filas) == 1)
    comprobar("CSV cp1252: la enye sobrevive",
              filas[0].datos["nombre"] == "RB-Niñez"
              and filas[0].datos["empresa"] == "Compañía Andina de Nuevas Redes")
    comprobar("CSV cp1252: avisa de que asumio la codificacion",
              any("cp1252" in a for a in avisos))

    # 5. Cabecera en otro orden y con mayusculas/espacios: la gente reordena
    # las columnas al copiar y pegar de su propia hoja.
    otro_orden = (
        "IP , Grupo ,NOMBRE,puerto,Empresa\r\n"
        "10.50.0.1,bts,BTS-Zamora,22,Austro Net\r\n"
    ).encode("utf-8")
    filas, avisos = leer(otro_orden, "equipos.csv")
    comprobar("cabecera en otro orden: mapea por nombre, no por posicion",
              filas[0].datos == {"nombre": "BTS-Zamora", "empresa": "Austro Net",
                                 "ip": "10.50.0.1", "puerto": "22", "grupo": "bts",
                                 "intervalo_minutos": "", "usuario": "",
                                 "clave": ""})
    comprobar("cabecera en otro orden: no se queja de nada", avisos == [])

    # 6. Sinonimos: nombres que la gente escribe de verdad.
    sinonimos = (
        "Router;Cliente;Direccion IP;Puerto SSH;Zona;Observaciones\r\n"
        "CPE-Yantzaza;Austro Net;10.60.0.7;22;cpes;revisar antena\r\n"
    ).encode("utf-8")
    filas, avisos = leer(sinonimos, "equipos.csv")
    comprobar("sinonimos: router/cliente/direccion ip/puerto ssh/zona valen",
              filas[0].datos == {"nombre": "CPE-Yantzaza", "empresa": "Austro Net",
                                 "ip": "10.60.0.7", "puerto": "22", "grupo": "cpes",
                                 "intervalo_minutos": "", "usuario": "",
                                 "clave": ""})
    comprobar("sinonimos: avisa de la columna que no entiende",
              any("Observaciones" in a for a in avisos))

    # 6 bis. Sinonimos del intervalo: nadie escribe 'intervalo_minutos' a mano.
    # Cada cabecera se prueba sola porque todas apuntan a la misma columna y
    # juntas se pisarian entre si.
    for cabecera, valor in (
        ("Intervalo", "1440"),
        ("Intervalo Minutos", "60"),
        ("intervalo_minutos", "30"),
        ("Cada cuanto", "120"),
        ("Frecuencia", "480"),
        ("MINUTOS", "15"),
    ):
        crudo = (
            f"nombre,ip,{cabecera}\r\n"
            f"R-Intervalo,10.61.0.1,{valor}\r\n"
        ).encode("utf-8")
        filas, avisos = leer(crudo, "equipos.csv")
        comprobar(f"sinonimo de intervalo '{cabecera}': se reconoce ({avisos})",
                  avisos == []
                  and filas[0].datos["intervalo_minutos"] == valor)

    # Y con el intervalo vacio, que es el caso normal: hereda el global.
    intervalo_vacio = (
        "nombre,ip,intervalo\r\n"
        "R-Hereda,10.61.0.2,\r\n"
    ).encode("utf-8")
    filas, avisos = leer(intervalo_vacio, "equipos.csv")
    comprobar("intervalo vacio: llega vacio y no se inventa un numero",
              filas[0].datos["intervalo_minutos"] == "")
    equipo, errores = validar_equipo(
        "R-Hereda", "Austro Net", "10.61.0.2", "", "cpes", existentes=[],
        intervalo=filas[0].datos["intervalo_minutos"],
    )
    comprobar(f"intervalo vacio: pasa validar_equipo como 0 ({errores})",
              equipo is not None and equipo.intervalo_minutos == 0)

    # 6 ter. Sinonimos de las credenciales por equipo. Cada pareja se prueba
    # sola, igual que el intervalo: varias cabeceras apuntan a la misma columna.
    for cabecera_usuario, cabecera_clave in (
        ("usuario", "clave"),
        ("Usuario SSH", "Clave SSH"),
        ("User", "Password"),
        ("Login", "Contraseña"),
        ("usuario", "Pass"),
    ):
        crudo = (
            f"nombre,ip,{cabecera_usuario},{cabecera_clave}\r\n"
            "R-Propias,10.62.0.1,soporte,s3cr3ta\r\n"
        ).encode("utf-8")
        filas, avisos = leer(crudo, "equipos.csv")
        comprobar(
            f"sinonimos de credencial '{cabecera_usuario}'/'{cabecera_clave}': "
            f"se reconocen ({avisos})",
            avisos == []
            and filas[0].datos["usuario"] == "soporte"
            and filas[0].datos["clave"] == "s3cr3ta",
        )

    # Y el caso normal: columnas presentes pero vacias, que es "hereda las
    # credenciales generales de config.yaml".
    credenciales_vacias = (
        "nombre,ip,usuario,clave\r\n"
        "R-Hereda,10.62.0.2,,\r\n"
    ).encode("utf-8")
    filas, avisos = leer(credenciales_vacias, "equipos.csv")
    comprobar("credenciales vacias: llegan vacias y no se inventa nada",
              avisos == [] and filas[0].datos["usuario"] == ""
              and filas[0].datos["clave"] == "")
    equipo, errores = validar_equipo(
        "R-Hereda", "Austro Net", "10.62.0.2", "", "cpes", existentes=[],
        usuario=filas[0].datos["usuario"], clave=filas[0].datos["clave"],
    )
    comprobar(f"credenciales vacias: pasan validar_equipo sin error ({errores})",
              equipo is not None and equipo.usuario == "" and equipo.clave == "")

    # Solo una de las dos: el importador no opina (no es problema del ARCHIVO)
    # y validar_equipo tampoco lo rechaza; el aviso de la mezcla lo da
    # inventory.cargar al releer el inventario ya guardado.
    solo_usuario = (
        "nombre,ip,usuario,clave\r\n"
        "R-Manco,10.62.0.3,soporte,\r\n"
    ).encode("utf-8")
    filas, avisos = leer(solo_usuario, "equipos.csv")
    comprobar("solo usuario: el importador lo deja pasar sin quejarse",
              avisos == [] and filas[0].datos["usuario"] == "soporte"
              and filas[0].datos["clave"] == "")
    equipo, errores = validar_equipo(
        "R-Manco", "Austro Net", "10.62.0.3", "", "cpes", existentes=[],
        usuario=filas[0].datos["usuario"], clave=filas[0].datos["clave"],
    )
    comprobar(f"solo usuario: validar_equipo tampoco lo rechaza ({errores})",
              equipo is not None and equipo.usuario == "soporte")

    # 7. Sin columna ip no hay nada que hacer: es la unica imprescindible.
    sin_ip = b"nombre,empresa,puerto,grupo\r\nBTS-X,Andinanet,22,bts\r\n"
    lanza("sin columna ip: lanza ErrorImportacion", lambda: leer(sin_ip, "x.csv"))
    try:
        leer(sin_ip, "x.csv")
    except ErrorImportacion as exc:
        mensaje = str(exc)
        comprobar("sin columna ip: el mensaje dice que se esperaba",
                  "ip" in mensaje and "nombre" in mensaje)
        comprobar("sin columna ip: el mensaje dice que se encontro",
                  "puerto" in mensaje and "grupo" in mensaje)

    # 8. Filas vacias intercaladas: quedan al borrar el contenido de una fila
    # en vez de la fila entera. No son un error y no deben desplazar nada.
    con_huecos = (
        "nombre,empresa,ip,puerto,grupo\r\n"
        "R1,Andinanet,10.70.0.1,22,core\r\n"
        ",,,,\r\n"
        "\r\n"
        "R2,Andinanet,10.70.0.2,22,core\r\n"
        ",,,,\r\n"
    ).encode("utf-8")
    filas, avisos = leer(con_huecos, "equipos.csv")
    comprobar("filas vacias: se saltan sin quejarse",
              len(filas) == 2 and avisos == [])
    comprobar("filas vacias: el numero de fila sigue siendo el del archivo",
              [f.numero for f in filas] == [2, 5])

    # 9. Nombre vacio a proposito: no es un error de este modulo. Se le
    # preguntara su identidad al router al darlo de alta.
    sin_nombre = (
        "empresa,ip,puerto,grupo\r\n"
        "Andinanet,10.80.0.1,,\r\n"
    ).encode("utf-8")
    filas, avisos = leer(sin_nombre, "equipos.csv")
    comprobar("sin columna nombre: se importa igual, el nombre queda vacio",
              len(filas) == 1 and filas[0].datos["nombre"] == "")

    # 9 bis. Datos mas alla de la ultima columna declarada: alguien escribe una
    # nota a la derecha de la tabla sin ponerle cabecera. Se ignora, pero hay
    # que decirlo o parecera que el importador se comio un dato.
    de_sobra = (
        "nombre,empresa,ip\r\n"
        "R1,Andinanet,10.85.0.1\r\n"
        "R2,Andinanet,10.85.0.2,revisar\r\n"
    ).encode("utf-8")
    filas, avisos = leer(de_sobra, "equipos.csv")
    comprobar("datos fuera de la cabecera: no descartan la fila", len(filas) == 2)
    comprobar("datos fuera de la cabecera: avisa una sola vez y dice la fila",
              len(avisos) == 1 and "fila 3" in avisos[0])

    # 10. Ciclo completo: la plantilla que se descarga tiene que poder subirse.
    # Si esto falla, el primero que se estrella es quien la usa tal cual.
    plantilla = plantilla_csv()
    filas, avisos = leer(plantilla, "plantilla.csv")
    comprobar("plantilla CSV: se puede volver a importar", len(filas) >= 2)
    comprobar("plantilla CSV: no produce avisos al leerse", avisos == [])
    comprobar("plantilla CSV: trae un ejemplo con el nombre vacio",
              any(f.datos["nombre"] == "" for f in filas))
    comprobar("plantilla CSV: lleva BOM para que Excel respete las tildes",
              plantilla.startswith(b"\xef\xbb\xbf"))
    comprobar("plantilla CSV: trae la columna intervalo_minutos",
              b"intervalo_minutos" in plantilla)
    comprobar("plantilla CSV: un ejemplo deja el intervalo vacio (hereda)",
              any(f.datos["intervalo_minutos"] == "" for f in filas))
    comprobar("plantilla CSV: otro ejemplo muestra un intervalo alto",
              any(f.datos["intervalo_minutos"] == "1440" for f in filas))
    comprobar("plantilla CSV: trae las columnas usuario y clave",
              b"usuario" in plantilla and b"clave" in plantilla)
    comprobar("plantilla CSV: casi todos los ejemplos dejan las credenciales "
              "vacias (el caso normal es heredarlas)",
              len([f for f in filas
                   if not f.datos["usuario"] and not f.datos["clave"]]) >= 2)
    con_credenciales = [f for f in filas if f.datos["usuario"]]
    comprobar("plantilla CSV: un unico ejemplo lleva credenciales propias",
              len(con_credenciales) == 1
              and con_credenciales[0].datos["clave"] != "")
    comprobar("plantilla CSV: las credenciales de ejemplo son obviamente falsas",
              con_credenciales[0].datos["usuario"] == "usuario-de-ejemplo"
              and con_credenciales[0].datos["clave"] == "cambiame")

    # Y sus filas de ejemplo tienen que pasar la validacion real: una plantilla
    # cuyos propios ejemplos no se pueden dar de alta es una trampa.
    nombradas = [f for f in filas if f.datos["nombre"]]
    equipo, errores = validar_equipo(
        nombradas[0].datos["nombre"], nombradas[0].datos["empresa"],
        nombradas[0].datos["ip"], nombradas[0].datos["puerto"],
        nombradas[0].datos["grupo"], existentes=[],
        intervalo=nombradas[0].datos["intervalo_minutos"],
    )
    comprobar(f"plantilla CSV: los ejemplos pasan validar_equipo ({errores})",
              equipo is not None)

    con_intervalo = [f for f in nombradas if f.datos["intervalo_minutos"]][0]
    equipo, errores = validar_equipo(
        con_intervalo.datos["nombre"], con_intervalo.datos["empresa"],
        con_intervalo.datos["ip"], con_intervalo.datos["puerto"],
        con_intervalo.datos["grupo"], existentes=[],
        intervalo=con_intervalo.datos["intervalo_minutos"],
        usuario=con_intervalo.datos["usuario"],
        clave=con_intervalo.datos["clave"],
    )
    comprobar(f"plantilla CSV: el ejemplo con intervalo valida ({errores})",
              equipo is not None and equipo.intervalo_minutos == 1440)
    comprobar("plantilla CSV: el ejemplo con credenciales llega entero a "
              "validar_equipo",
              equipo is not None and equipo.usuario == "usuario-de-ejemplo"
              and equipo.clave == "cambiame")

    # 11. Extensiones que no se admiten.
    lanza("extension desconocida: lanza", lambda: leer(b"x", "equipos.txt"))
    lanza("xls antiguo: lanza con mensaje propio", lambda: leer(b"x", "equipos.xls"))
    lanza("archivo vacio: lanza", lambda: leer(b"", "equipos.csv"))

    # 12. XLSX. Solo si openpyxl esta instalado: es una dependencia OPCIONAL y
    # el resto del sistema (incluido el respaldo) funciona sin ella.
    comprobar("hay_soporte_xlsx coincide con lo que se puede importar",
              hay_soporte_xlsx() == _openpyxl_disponible())

    if not hay_soporte_xlsx():
        print("\n(openpyxl no instalado: se omiten las pruebas de .xlsx)")
        lanza("sin openpyxl: leer un .xlsx lanza",
              lambda: leer(b"PK\x03\x04", "equipos.xlsx"))
        lanza("sin openpyxl: plantilla_xlsx lanza", plantilla_xlsx)
    else:
        import openpyxl

        # Ciclo plantilla -> leer, igual que en CSV.
        libro_bytes = plantilla_xlsx()
        filas, avisos = leer(libro_bytes, "plantilla.xlsx")
        comprobar("plantilla XLSX: se puede volver a importar", len(filas) >= 2)
        comprobar("plantilla XLSX: no produce avisos al leerse", avisos == [])
        comprobar("plantilla XLSX: trae un ejemplo con el nombre vacio",
                  any(f.datos["nombre"] == "" for f in filas))

        libro = openpyxl.load_workbook(io.BytesIO(libro_bytes))
        hoja = libro["Equipos"]
        comprobar("plantilla XLSX: cabecera en negrita", hoja["A1"].font.bold)
        comprobar("plantilla XLSX: cabecera inmovilizada",
                  hoja.freeze_panes == "A2")
        comprobar("plantilla XLSX: columnas con ancho suficiente",
                  hoja.column_dimensions["B"].width >= 20)
        comprobar("plantilla XLSX: las notas van en otra hoja",
                  "Instrucciones" in libro.sheetnames)
        comprobar("plantilla XLSX: las notas explican el nombre vacio",
                  any("identity" in str(c[0].value or "")
                      for c in libro["Instrucciones"].iter_rows()))
        cabecera_xlsx = [c.value for c in hoja[1]]
        comprobar("plantilla XLSX: la cabecera trae intervalo_minutos",
                  "intervalo_minutos" in cabecera_xlsx)
        comprobar("plantilla XLSX: la cabecera termina en usuario,clave",
                  cabecera_xlsx[-2:] == ["usuario", "clave"])
        comprobar("plantilla XLSX: las columnas nuevas tienen ancho suficiente",
                  hoja.column_dimensions["G"].width >= 18
                  and hoja.column_dimensions["H"].width >= 18)
        instrucciones = "\n".join(
            str(c[0].value or "") for c in libro["Instrucciones"].iter_rows()
        )
        comprobar("plantilla XLSX: las notas dicen que vacio = intervalo general",
                  "VACIO" in instrucciones and "intervalo general" in instrucciones)
        comprobar("plantilla XLSX: las notas explican que usuario/clave se heredan",
                  "usuario general" in instrucciones)
        comprobar("plantilla XLSX: las notas avisan de las contrasenas en claro",
                  "CONTRASENAS EN CLARO" in instrucciones)
        comprobar("plantilla XLSX: las notas dicen que no se mande por correo "
                  "y que se borre",
                  "correo" in instrucciones and "BORRALO" in instrucciones)
        libro.close()

        comprobar("plantilla XLSX: un ejemplo deja el intervalo vacio (hereda)",
                  any(f.datos["intervalo_minutos"] == "" for f in filas))
        comprobar("plantilla XLSX: otro ejemplo muestra un intervalo alto",
                  any(f.datos["intervalo_minutos"] == "1440" for f in filas))
        comprobar("plantilla XLSX: los ejemplos dejan las credenciales vacias "
                  "salvo uno",
                  len([f for f in filas if f.datos["usuario"]]) == 1)
        comprobar("plantilla XLSX: el ejemplo con credenciales es falso a la vista",
                  any(f.datos["usuario"] == "usuario-de-ejemplo"
                      and f.datos["clave"] == "cambiame" for f in filas))

        # Puerto como NUMERO: es lo que guarda Excel si nadie formatea la
        # columna como texto, y openpyxl lo devuelve como int o como float.
        # str(22.0) daria "22.0" y ningun puerto valido.
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.append(["nombre", "empresa", "ip", "puerto", "grupo", "intervalo",
                     "User", "Password"])
        hoja.append(["R-Entero", "Andinanet", "10.90.0.1", 22, "core", 60,
                     "soporte", "s3cr3ta"])
        hoja.append(["R-Float", "Andinanet", "10.90.0.2", 2222.0, "core", 1440.0,
                     None, None])
        hoja.append([None] * 8)  # fila vacia intercalada
        hoja.append(["R-Ultimo", "Andinanet", "10.90.0.3", None, "core", None,
                     None, None])
        crudo = io.BytesIO()
        libro.save(crudo)
        libro.close()

        filas, avisos = leer(crudo.getvalue(), "equipos.xlsx")
        comprobar("XLSX: salta la fila vacia intercalada", len(filas) == 3)
        comprobar("XLSX: el puerto entero llega como '22'",
                  filas[0].datos["puerto"] == "22")
        comprobar("XLSX: el puerto float NO llega como '2222.0'",
                  filas[1].datos["puerto"] == "2222")
        comprobar("XLSX: la celda vacia queda como cadena vacia",
                  filas[2].datos["puerto"] == "")
        comprobar("XLSX: el numero de fila es el de Excel",
                  [f.numero for f in filas] == [2, 3, 5])

        comprobar("XLSX: el intervalo entero llega como '60'",
                  filas[0].datos["intervalo_minutos"] == "60")
        comprobar("XLSX: el intervalo float NO llega como '1440.0'",
                  filas[1].datos["intervalo_minutos"] == "1440")
        comprobar("XLSX: el intervalo vacio queda como cadena vacia",
                  filas[2].datos["intervalo_minutos"] == "")

        comprobar("XLSX: los sinonimos User/Password tambien valen en Excel",
                  filas[0].datos["usuario"] == "soporte"
                  and filas[0].datos["clave"] == "s3cr3ta")
        comprobar("XLSX: las celdas de credenciales vacias quedan en cadena vacia",
                  filas[1].datos["usuario"] == ""
                  and filas[1].datos["clave"] == "")

        equipo, errores = validar_equipo(
            filas[1].datos["nombre"], filas[1].datos["empresa"],
            filas[1].datos["ip"], filas[1].datos["puerto"],
            filas[1].datos["grupo"], existentes=[],
            intervalo=filas[1].datos["intervalo_minutos"],
        )
        comprobar(f"XLSX: el puerto numerico pasa validar_equipo ({errores})",
                  equipo is not None and equipo.puerto == 2222)
        comprobar("XLSX: el intervalo numerico llega entero a validar_equipo",
                  equipo is not None and equipo.intervalo_minutos == 1440)

        # Un .xlsx que no lo es (alguien renombra un CSV) no debe reventar con
        # una traza de zipfile en la cara del que sube el archivo.
        lanza("XLSX corrupto: lanza ErrorImportacion",
              lambda: leer(b"esto no es un zip", "equipos.xlsx"))

    print()
    if FALLOS:
        print(f"{len(FALLOS)} prueba(s) fallidas:")
        for f in FALLOS:
            print(f"  - {f}")
        raise SystemExit(1)
    print("Todas las pruebas pasaron.")


def _openpyxl_disponible() -> bool:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


if __name__ == "__main__":
    main()
