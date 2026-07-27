"""Alta masiva de equipos desde una plantilla de Excel o CSV.

El panel web ofrece "descargar plantilla" y "subir archivo": alguien llena la
hoja con los routers de un cliente y los da de alta de golpe, en vez de teclear
trescientas veces el mismo formulario.

Este modulo SOLO extrae filas crudas del archivo. No valida los datos de cada
equipo: de eso se encarga inventory.validar_equipo, que ya sabe que es un
nombre aceptable, que IP tiene sentido y que nombres estan repetidos. Duplicar
esas reglas aqui garantizaria que un dia dejen de coincidir y que el alta por
Excel acepte lo que el formulario rechaza.

Reparto de responsabilidades:

  - leer() devuelve (filas, avisos). Los AVISOS son problemas del ARCHIVO:
    cabecera desconocida, columnas de sobra, hoja sin datos, codificacion
    dudosa. Los errores de cada equipo los dira validar_equipo despues.
  - ErrorImportacion es lo que impide siquiera empezar: formato no soportado,
    archivo vacio o ilegible, falta la columna 'ip'.

Por que 'nombre' puede venir VACIO a proposito: si la plantilla no lo trae, el
sistema le pregunta al propio router su /system identity al darlo de alta. Quien
llena la hoja tiene a mano la lista de IPs del cliente, no los nombres exactos
de cada equipo, y obligarle a inventarselos solo consigue inventarios con
nombres que no coinciden con la realidad. La unica columna imprescindible es
'ip'.

Sinonimos aceptados en la cabecera (sin distinguir mayusculas, tildes,
guiones bajos ni espacios sobrantes):

  nombre             <- nombre, router, equipo, identidad, nombre del equipo
  empresa            <- empresa, cliente, razon social
  ip                 <- ip, direccion, direccion ip, host, ip o dns
  puerto             <- puerto, puerto ssh, port
  grupo              <- grupo, zona, categoria
  intervalo_minutos  <- intervalo, intervalo minutos, cada cuanto, frecuencia,
                        minutos
  usuario            <- usuario, usuario ssh, user, login
  clave              <- clave, clave ssh, password, contrasena, pass

OJO con las columnas 'usuario' y 'clave': son opcionales (vacias = se usan las
credenciales generales de config.yaml), pero si alguien las llena, el archivo
que se sube pasa a llevar CONTRASENAS EN CLARO. La plantilla lo advierte en sus
notas: ese archivo no se manda por correo ni se deja en la carpeta de descargas,
se borra en cuanto se ha importado.

El soporte de .xlsx depende de openpyxl, que es una dependencia OPCIONAL (ver
requirements.txt): el respaldo desatendido de madrugada no puede quedarse sin
correr porque falte una libreria que solo hace falta para importar una hoja de
calculo desde el navegador. Por eso openpyxl se importa DENTRO de las funciones
que lo necesitan y nunca en la cabecera del modulo.
"""

from __future__ import annotations

import csv
import io
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

# Columnas que entiende el sistema, en el orden en que se escriben la plantilla
# y el inventario (ver inventory.CABECERA).
COLUMNAS = (
    "nombre",
    "empresa",
    "ip",
    "puerto",
    "grupo",
    "intervalo_minutos",
    "usuario",
    "clave",
)

# Cabecera -> columna real. Las claves ya vienen normalizadas (_normalizar):
# minusculas, sin tildes y con un solo espacio entre palabras. Son los nombres
# que la gente escribe de verdad cuando rehace la plantilla a su manera.
SINONIMOS = {
    "nombre": "nombre",
    "router": "nombre",
    "equipo": "nombre",
    "identidad": "nombre",
    "nombre del equipo": "nombre",
    "empresa": "empresa",
    "cliente": "empresa",
    "razon social": "empresa",
    "ip": "ip",
    "direccion": "ip",
    "direccion ip": "ip",
    "host": "ip",
    "ip o dns": "ip",
    "puerto": "puerto",
    "puerto ssh": "puerto",
    "port": "puerto",
    "grupo": "grupo",
    "zona": "grupo",
    "categoria": "grupo",
    # 'intervalo_minutos' se normaliza a 'intervalo minutos' (_normalizar
    # cambia el guion bajo por espacio), asi que la cabecera canonica ya entra
    # por esa clave; las demas son como lo escribe la gente en su propia hoja.
    "intervalo": "intervalo_minutos",
    "intervalo minutos": "intervalo_minutos",
    "cada cuanto": "intervalo_minutos",
    "frecuencia": "intervalo_minutos",
    "minutos": "intervalo_minutos",
    # Credenciales propias del equipo. 'contrasena' entra normalizada: la enye
    # y la tilde se pierden en _normalizar, asi que "Contraseña" cae aqui.
    "usuario": "usuario",
    "usuario ssh": "usuario",
    "user": "usuario",
    "login": "usuario",
    "clave": "clave",
    "clave ssh": "clave",
    "password": "clave",
    "contrasena": "clave",
    "pass": "clave",
}

# Tope de filas de datos. Esto lo sube alguien por HTTP: sin limite, un archivo
# de un millon de filas tiene al proceso web ocupado y en memoria un buen rato.
# 5000 equipos es varias veces el inventario mas grande que se espera.
MAX_FILAS = 5000

EXTENSIONES = (".csv", ".xlsx")

# Ejemplos de la plantilla. Realistas a proposito (un ISP con torres y nodos de
# fibra) para que se vea como se llena, e incluyen a proposito una fila SIN
# nombre: es la forma de anunciar que ese campo se puede dejar en blanco.
#
# El intervalo se muestra igual: vacio en las dos primeras (heredan el
# intervalo general) y 1440 en la ultima, un equipo que casi no cambia y al que
# no tiene sentido preguntarle mas de una vez al dia.
#
# Usuario y clave siguen el mismo criterio: vacios en las dos primeras (el caso
# normal, se usan los de config.yaml) y llenos SOLO en la ultima, para que se
# vea donde van sin sugerir que hay que llenarlos siempre. Los valores son
# obviamente de mentira a proposito: nadie debe poder copiar la plantilla y
# creer que ahi habia una credencial de verdad.
EJEMPLOS = (
    ("BTS-Cuenca-01", "Andinanet S.A.", "10.20.1.1", "22", "bts", "", "", ""),
    ("", "Andinanet S.A.", "10.20.1.2", "", "bts", "", "", ""),
    (
        "NODO-Loja-Core",
        "Fibra Austral Cia. Ltda.",
        # Rango reservado por la RFC 5737 para documentacion: no existe en
        # internet. En un ejemplo publico, una IP con pinta de real acaba
        # apuntando a la red de alguien.
        "203.0.113.10",
        "2222",
        "core",
        "1440",
        "usuario-de-ejemplo",
        "cambiame",
    ),
)

NOTAS = (
    "Como llenar esta plantilla",
    "",
    "ip       OBLIGATORIA. IP o nombre DNS por el que se llega al equipo.",
    "nombre   Opcional. Si se deja VACIO, al dar de alta el equipo el sistema",
    "         le pregunta su /system identity y usa ese nombre.",
    "empresa  Nombre del cliente. Agrupa los respaldos en su propia carpeta.",
    "puerto   Opcional, por defecto 22 (SSH). Ojo: 8291 es Winbox y 8728 la",
    "         API; por ahi el respaldo no responde.",
    "grupo    Opcional, por defecto 'mikrotiks'. Sirve para ordenar la lista.",
    "intervalo_minutos",
    "         Opcional. Cada cuantos minutos se consulta ESE equipo. Si se",
    "         deja VACIO se usa el intervalo general del sistema, que es lo",
    "         normal. Ponle un numero solo si ese equipo necesita otro ritmo:",
    "         por ejemplo 1440 (una vez al dia) en un router que casi no",
    "         cambia. Por debajo de 5 minutos se avisa: cada respaldo abre una",
    "         sesion SSH y pide un export completo.",
    "usuario  Opcional. Usuario SSH de ESE equipo. Si se deja VACIO se usa el",
    "         usuario general del sistema, que es lo normal.",
    "clave    Opcional. Contrasena SSH de ESE equipo. Vacia = la general.",
    "",
    "AVISO: si rellenas 'usuario' y 'clave', este archivo pasa a contener",
    "CONTRASENAS EN CLARO de tus routers. Tratalo como tal: no lo mandes por",
    "correo ni por chat, no lo dejes en una carpeta compartida ni en Descargas,",
    "y BORRALO en cuanto lo hayas importado. Rellenalas solo en los equipos que",
    "no usan las credenciales generales; en el resto, dejalas vacias.",
    "Llenar solo una de las dos se avisa al cargar el inventario: se mezclaria",
    "con la general, y casi siempre es un descuido.",
    "",
    "No hace falta respetar el orden de las columnas ni las mayusculas.",
    "Las filas de ejemplo se pueden borrar o dejar: revisalas antes de subir.",
    "Las credenciales de la fila de ejemplo son falsas: no sirven para nada.",
)


class ErrorImportacion(Exception):
    """El archivo no se pudo leer: formato, cabecera o contenido inservibles."""


@dataclass
class Fila:
    """Una fila de datos tal cual venia, sin validar.

    'numero' es la fila REAL del archivo (la cabecera suele ser la 1), para
    poder decirle a quien la subio "fila 7" y que la encuentre a la primera.
    """

    numero: int
    datos: dict[str, str] = field(default_factory=dict)


def hay_soporte_xlsx() -> bool:
    """True si openpyxl esta instalado y por tanto se pueden leer .xlsx."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def _falta_openpyxl() -> ErrorImportacion:
    return ErrorImportacion(
        "falta openpyxl y sin el no se pueden leer archivos .xlsx; "
        "instalalo con 'pip install openpyxl' o sube el archivo en CSV "
        "(Excel: Guardar como > CSV)"
    )


def _normalizar(texto: str) -> str:
    """Deja una cabecera comparable: minusculas, sin tildes, un solo espacio.

    Excel devuelve 'Direccion IP ', alguien escribe 'DIRECCION_IP' y otro
    'Dirección IP'. Las tres cosas son la misma columna y no vale la pena
    fallar por eso.
    """
    base = unicodedata.normalize("NFKD", (texto or "").strip())
    base = "".join(c for c in base if not unicodedata.combining(c))
    base = base.lower().replace("_", " ").replace("-", " ")
    return " ".join(base.split())


def _mapear_cabecera(celdas: list[str]) -> tuple[dict[str, int], list[str]]:
    """Cabecera -> {columna: indice}. Acepta cualquier orden. Devuelve avisos."""
    indices: dict[str, int] = {}
    avisos: list[str] = []
    desconocidas: list[str] = []

    for i, celda in enumerate(celdas):
        clave = _normalizar(celda)
        if not clave:
            continue  # celda de cabecera vacia: Excel las deja de sobra
        columna = SINONIMOS.get(clave)
        if columna is None:
            desconocidas.append(celda.strip())
            continue
        if columna in indices:
            # Dos cabeceras distintas apuntando a lo mismo ('ip' y 'host').
            avisos.append(
                f"la columna '{columna}' viene mas de una vez; se usa la primera"
            )
            continue
        indices[columna] = i

    if desconocidas:
        avisos.append(
            "columnas que no se entienden y se ignoran: "
            + ", ".join(f"'{c}'" for c in desconocidas)
        )

    # Sin ip no hay a donde conectarse: lo demas tiene valor por defecto o se
    # le pregunta al equipo, esto no.
    if "ip" not in indices:
        encontradas = [c.strip() for c in celdas if c.strip()]
        raise ErrorImportacion(
            "falta la columna obligatoria 'ip'. Se esperaba una cabecera con: "
            + ", ".join(COLUMNAS)
            + " (solo 'ip' es obligatoria). Se encontro: "
            + (", ".join(encontradas) if encontradas else "(cabecera vacia)")
        )

    return indices, avisos


def _extraer(filas, origen: str) -> tuple[list[Fila], list[str]]:
    """Nucleo comun a CSV y XLSX: recibe (numero, celdas) ya en texto."""
    avisos: list[str] = []
    indices: dict[str, int] | None = None
    resultado: list[Fila] = []
    leidas = 0
    ancho_cabecera = 0
    sobra_avisado = False

    for numero, celdas in filas:
        # Fila completamente vacia: en Excel siempre sobran al final, y en
        # medio aparecen cuando alguien borra el contenido en vez de la fila.
        if not any(celdas):
            continue

        if indices is None:
            indices, avisos_cabecera = _mapear_cabecera(celdas)
            avisos.extend(avisos_cabecera)
            # Se recorta a la ultima celda con texto: Excel puede devolver una
            # cabecera con diez celdas vacias detras y eso no es "de sobra".
            ancho_cabecera = max(
                (i + 1 for i, c in enumerate(celdas) if c), default=0
            )
            continue

        leidas += 1
        if leidas > MAX_FILAS:
            avisos.append(
                f"el archivo trae mas de {MAX_FILAS} filas; se leyeron las "
                f"primeras {MAX_FILAS} y el resto se ignoro"
            )
            break

        # Se avisa una sola vez: si a una fila le sobra una columna, le sobra a
        # las 300 y repetirlo enterraria el resto de avisos.
        utiles = max((i + 1 for i, c in enumerate(celdas) if c), default=0)
        if utiles > ancho_cabecera and not sobra_avisado:
            sobra_avisado = True
            avisos.append(
                f"fila {numero}: hay datos en columnas que la cabecera no "
                "declara; se ignoran"
            )

        datos = {
            columna: (celdas[i].strip() if i < len(celdas) else "")
            for columna, i in indices.items()
        }
        # Todas las claves de COLUMNAS siempre presentes: quien consuma esto
        # no tiene que andar con .get() y valores por defecto.
        resultado.append(
            Fila(numero=numero, datos={c: datos.get(c, "") for c in COLUMNAS})
        )

    if indices is None:
        # Ni siquiera hubo cabecera: no es un aviso, es que no hay nada que leer.
        raise ErrorImportacion(f"{origen} esta vacio: no tiene ni cabecera")

    if not resultado:
        avisos.append(f"{origen} solo trae la cabecera: no hay ningun equipo")

    return resultado, avisos


# --- CSV --------------------------------------------------------------------


def _decodificar_csv(contenido: bytes) -> tuple[str, list[str]]:
    """Bytes -> texto. utf-8 (con o sin BOM) y, si no, cp1252."""
    avisos: list[str] = []
    try:
        # utf-8-sig se come el BOM que Excel mete al guardar como CSV UTF-8.
        return contenido.decode("utf-8-sig"), avisos
    except UnicodeDecodeError:
        pass

    try:
        # "CSV (delimitado por comas)" de Excel en Windows guarda en la pagina
        # de codigos local, no en UTF-8: una enye llega como un solo byte 0xF1.
        texto = contenido.decode("cp1252")
    except UnicodeDecodeError as exc:
        raise ErrorImportacion(
            "el archivo no es texto valido (ni UTF-8 ni cp1252); "
            "vuelve a guardarlo desde Excel como CSV UTF-8"
        ) from exc

    avisos.append(
        "el archivo no estaba en UTF-8: se asumio cp1252 (Excel en Windows). "
        "Revisa que las tildes y las enyes se vean bien"
    )
    return texto, avisos


def _separador(texto: str) -> str:
    """Coma, punto y coma o tabulador: lo que use el Excel de quien lo guardo."""
    muestra = texto[:4096]
    try:
        return csv.Sniffer().sniff(muestra, delimiters=",;\t").delimiter
    except csv.Error:
        pass

    # El Sniffer se rinde con archivos de una sola columna o de pocas lineas,
    # que es justo el caso de una plantilla con dos routers. Se prueba a mano:
    # coma primero (el formato canonico del inventario) y punto y coma despues
    # (lo que produce Excel en configuracion regional espanola).
    primera = texto.splitlines()[0] if texto.splitlines() else ""
    for candidato in (",", ";", "\t"):
        if candidato in primera:
            return candidato
    return ","


def _leer_csv(contenido: bytes) -> tuple[list[Fila], list[str]]:
    texto, avisos = _decodificar_csv(contenido)
    if not texto.strip():
        raise ErrorImportacion("el archivo CSV esta vacio")

    # newline="" lo exige el modulo csv; ademas asi el CRLF de Windows no deja
    # un retorno de carro pegado al final de la ultima columna.
    lector = csv.reader(io.StringIO(texto, newline=""), delimiter=_separador(texto))
    # El numero de fila se lleva a mano en vez de con lector.line_num porque
    # este ultimo cuenta las lineas fisicas y un campo entrecomillado con salto
    # de linea desplazaria la cuenta respecto a lo que se ve en Excel.
    filas = ((n, [(c or "").strip() for c in fila]) for n, fila in enumerate(lector, 1))

    encontradas, avisos_extra = _extraer(filas, "el CSV")
    return encontradas, avisos + avisos_extra


# --- XLSX -------------------------------------------------------------------


def _texto_celda(valor) -> str:
    """Celda de openpyxl -> texto recortado.

    Cuidado con los numeros: Excel guarda el puerto 22 como numero, openpyxl lo
    devuelve como int o como float segun el formato de la celda, y str(22.0) da
    "22.0", que luego no es un puerto valido. Los enteros disfrazados de float
    se devuelven sin la parte decimal. Lo mismo vale para intervalo_minutos: un
    1440 tecleado en Excel puede llegar como 1440.0 y "1440.0" no es un entero.
    """
    if valor is None:
        return ""
    if isinstance(valor, bool):
        # Poco probable, pero True no es un dato util en ninguna columna.
        return ""
    if isinstance(valor, float):
        return str(int(valor)) if valor.is_integer() else str(valor)
    if isinstance(valor, int):
        return str(valor)
    return str(valor).strip()


def _leer_xlsx(contenido: bytes) -> tuple[list[Fila], list[str]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise _falta_openpyxl() from exc

    try:
        # read_only: no carga la hoja entera en memoria, que esto viene de una
        # subida HTTP. data_only: interesa el resultado de las formulas, no la
        # formula (alguien concatena la IP con CONCATENAR y quedaria "=A2&...").
        libro = openpyxl.load_workbook(
            io.BytesIO(contenido), read_only=True, data_only=True
        )
    except Exception as exc:  # BadZipFile, InvalidFileException, KeyError...
        raise ErrorImportacion(
            f"no se pudo abrir el archivo como Excel (.xlsx): {exc}. "
            "Si es un .xls antiguo, guardalo como .xlsx o como CSV"
        ) from exc

    try:
        # La hoja ACTIVA: la que la persona tenia delante al guardar, que es la
        # que cree haber subido aunque el libro tenga varias.
        hoja = libro.active
        if hoja is None:
            raise ErrorImportacion("el libro de Excel no tiene ninguna hoja")

        filas = (
            (n, [_texto_celda(c) for c in fila])
            for n, fila in enumerate(hoja.iter_rows(values_only=True), 1)
        )
        encontradas, avisos = _extraer(filas, f"la hoja '{hoja.title}'")
    finally:
        # read_only deja descriptores abiertos hasta que se cierra el libro.
        libro.close()

    return encontradas, avisos


# --- Entrada publica --------------------------------------------------------


def leer(contenido: bytes, nombre_archivo: str) -> tuple[list[Fila], list[str]]:
    """Extrae las filas de un archivo subido. Devuelve (filas, avisos).

    El formato se decide por la extension de 'nombre_archivo': el contenido
    llega en memoria desde una subida HTTP y no hay ruta en disco que mirar.

    Las filas NO estan validadas: pasalas por inventory.validar_equipo, que es
    quien decide si un equipo se puede dar de alta y con que mensajes de error.
    """
    extension = Path(nombre_archivo or "").suffix.lower()

    if extension == ".csv":
        return _leer_csv(contenido)
    if extension == ".xlsx":
        return _leer_xlsx(contenido)

    if extension in (".xls", ".xlsm", ".ods"):
        raise ErrorImportacion(
            f"el formato '{extension}' no se admite; abre el archivo y guardalo "
            "como .xlsx o como CSV"
        )
    raise ErrorImportacion(
        f"no se reconoce la extension del archivo '{nombre_archivo}'; "
        f"se admiten: {', '.join(EXTENSIONES)}"
    )


# --- Plantillas -------------------------------------------------------------


def plantilla_csv() -> bytes:
    """Plantilla en CSV, lista para descargar.

    Sin comentarios ni notas dentro del archivo: en CSV cualquier linea de
    ayuda seria una fila mas y el propio importador la leeria como un equipo
    roto. Las explicaciones van en la version .xlsx.
    """
    buffer = io.StringIO(newline="")
    escritor = csv.writer(buffer, lineterminator="\r\n")
    escritor.writerow(COLUMNAS)
    for ejemplo in EJEMPLOS:
        escritor.writerow(ejemplo)

    # Con BOM: Excel en Windows abre un CSV sin BOM en cp1252 y destroza las
    # tildes de los nombres de empresa. leer() se lo come sin quejarse.
    return buffer.getvalue().encode("utf-8-sig")


def plantilla_xlsx() -> bytes:
    """Plantilla en Excel: cabecera en negrita, fija, y una hoja de notas."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise _falta_openpyxl() from exc

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Equipos"

    hoja.append(list(COLUMNAS))
    for ejemplo in EJEMPLOS:
        hoja.append(list(ejemplo))

    negrita = Font(bold=True)
    fondo = PatternFill("solid", fgColor="DDEBF7")
    for celda in hoja[1]:
        celda.font = negrita
        celda.fill = fondo
        celda.alignment = Alignment(horizontal="center")

    # Ancho a ojo por columna: con el ancho por defecto, "Fibra Austral Cia.
    # Ltda." se ve cortada y parece que el dato esta mal.
    for columna, ancho in zip("ABCDEFGH", (24, 30, 20, 10, 16, 18, 20, 18)):
        hoja.column_dimensions[columna].width = ancho

    # La cabecera queda a la vista al bajar por una lista de 300 equipos: sin
    # esto, a la fila 40 ya nadie sabe que columna esta llenando.
    hoja.freeze_panes = "A2"

    # Las notas van en su propia hoja y NO en la de datos: cualquier texto
    # suelto junto a la tabla acabaria leido como una fila o como una columna
    # desconocida al volver a subir la plantilla.
    notas = libro.create_sheet("Instrucciones")
    for linea in NOTAS:
        notas.append([linea])
    notas["A1"].font = Font(bold=True, size=13)
    notas.column_dimensions["A"].width = 95

    salida = io.BytesIO()
    libro.save(salida)
    libro.close()
    return salida.getvalue()
